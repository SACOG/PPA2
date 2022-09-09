# --------------------------------
# Name: utils.py
# Purpose: Provides general PPA functions that are used throughout various PPA scripts and are not specific to any one PPA script
# NOTE:
#    This version, in the __init__ method of the Publish class, hard-codes in variables
#   that normally come from the params parameter script. This is because, for some reason
# # when pulled from the params file, the values "stuck" between runs and the wrong output tabs
#   in the output Excel spreadsheet would be colored.
# Author: Darren Conly
# Last Updated: 8/6/2020
# Updated by: <name>
# Copyright:   (c) SACOG
# Python Version: 3.x
# --------------------------------

import os
# import pdb
import sys
import datetime as dt
import time
import gc
import csv
import math
import shutil

import openpyxl
from openpyxl.drawing.image import Image
import xlwings as xw
import pandas as pd
import arcpy

import ppa_input_params as params


def trace():
    import traceback, inspect
    tb = sys.exc_info()[2]
    tbinfo = traceback.format_tb(tb)[0]
    # script name + line number
    line = tbinfo.split(", ")[1]
    filename = inspect.getfile(inspect.currentframe())
    # Get Python syntax error
    synerror = traceback.format_exc().splitlines()[-1]
    return line, filename, synerror

    
def remove_forbidden_chars(in_str):
    '''Replaces forbidden characters with acceptable characters'''
    repldict = {"&":'And','%':'pct','/':'-'}
    
    for old, new in repldict.items():
        if old in in_str:
            out_str = in_str.replace(old, new)
        else:
            out_str = in_str
    
    return out_str
    
    
def esri_field_exists(in_tbl, field_name):
    fields = [f.name for f in arcpy.ListFields(in_tbl)]
    if field_name in fields:
        return True
    else:
        return False


def esri_object_to_df(in_esri_obj, esri_obj_fields, index_field=None):
    '''converts esri gdb table, feature class, feature layer, or SHP to pandas dataframe'''
    data_rows = []
    with arcpy.da.SearchCursor(in_esri_obj, esri_obj_fields) as cur:
        for row in cur:
            out_row = list(row)
            data_rows.append(out_row)

    out_df = pd.DataFrame(data_rows, index=index_field, columns=esri_obj_fields)
    return out_df


def return_perf_outcomes_options(project_type):
    xlsx = params.type_template_dict[project_type]
    xlsx_path = os.path.join(params.template_dir, xlsx)
    
    wb = openpyxl.load_workbook(xlsx_path)  # openpyxl.load_workbook(xlsx_path, read_only=False, keep_vba=True)
    sheets = wb.sheetnames
    
    # values in this list will be the potential performance outcomes from which users can choose
    perf_outcomes = [s for s in sheets if s not in params.sheets_all_reports] 
    return perf_outcomes
    
    
def rename_dict_keys(dict_in, new_key_dict):
    '''if dict in = {0:1} and dict out supposed to be {'zero':1}, this function renames the key accordingly per
    the new_key_dict (which for this example would be {0:'zero'}'''
    dict_out = {}
    for k, v in new_key_dict.items():
        if k in list(dict_in.keys()):
            dict_out[v] = dict_in[k]
        else:
            dict_out[v] = 0
    return dict_out



def join_xl_import_template(template_xlsx, template_sheet, in_df, joincolidx=0):
    '''takes in import tab of destination Excel sheet, then left joins to desired output dataframe to ensure that
    output CSV has same rows every time, even if data frame that you're joining doesn't
    have all records'''
    df_template = pd.read_excel(template_xlsx, template_sheet)
    df_template = pd.DataFrame(df_template[df_template.columns[joincolidx]]) # get rid of all columns except for data items column
    df_template = df_template.set_index(df_template.columns[joincolidx]) # set data items column to be the index
    
    df_out = df_template.join(in_df) # left join the df from import sheet template to the df with data based on index values
    
    return df_out

def append_proj_to_master_fc(project_fc, proj_attributes_dict, master_fc):
    '''Takes project line and appends it to master line feature class with all lines users have entered'''
    arcpy.AddMessage("Archiving project line geometry...")
    #get geometry of user-drawn input line
    try:
        fld_shape = "SHAPE@"
        geoms = []
        with arcpy.da.SearchCursor(project_fc, fld_shape) as cur:
            for row in cur:
                geoms.append(row[0])
        
        #make row that will be inserted into master fc
        new_row = geoms + [v for k, v in proj_attributes_dict.items()]
        
        # use insert cursor to add in appropriate project name, etc.
        fields = [fld_shape] + list(proj_attributes_dict.keys())
        
        inscur = arcpy.da.InsertCursor(master_fc, fields)
        inscur.insertRow(new_row)
        
        del inscur
        
        t_returns = (params.msg_ok,)
    except:
        msg = trace()
        t_returns = (msg,)
    
    return t_returns


class Publish(object):
    def __init__(self, in_df, xl_template, import_tab, xl_out, project_fc, ptype, selecd_po_sheets=None, 
                 proj_name='UnnamedProject'):
        # params from input arguments
        self.in_df = in_df
        self.xl_template = xl_template
        self.import_tab = import_tab
        self.xl_out = xl_out
        self.selecd_po_sheets = selecd_po_sheets # 3 performance outcomes selected by user
        self.proj_name = proj_name
        self.project_fc = project_fc #remember, this is a feature set!

        #=====WORKAROUND - these are normally supposed to be imported from params.py file======
        '''Root issue replication:
        1 - run the tool, selecting any number of outcomes
        2 - re-run the tool, but with different outcomes selected.
        Expected result = otuput Excel file has the tabs colored and moved for the outcomes selected for each run
        Actual result = output Excel for second run has tabs colored for the outcomes selected for
            both runs, e.g., tabs selected for run two have outcome tabs colored for outcomes selected in
            run 1 as well.'''
        ptype_fwy = 'Freeway'
        ptype_arterial = 'Arterial or Transit Expansion'
        ptype_sgr = 'Complete Street or State of Good Repair'
        ptype_commdesign = "Community Design"
        
        xlsx_disclaimer_sheet = '0BUsingThisReport'
        xlsx_titlepg_sheet = '0ATitlePg'
        xlsx_socequity_sheet = '8SocioEconEquity'

        # regardless of which perf outcomes user selects, these tabs will be printed to
        # every PDF report for the selected project type.
        sheets_all_reports_workarnd = {ptype_arterial: [xlsx_titlepg_sheet, xlsx_disclaimer_sheet, xlsx_socequity_sheet],
                              ptype_sgr: [xlsx_titlepg_sheet, xlsx_disclaimer_sheet, xlsx_socequity_sheet],
                              ptype_commdesign: [xlsx_titlepg_sheet, xlsx_disclaimer_sheet],
                              ptype_fwy: [xlsx_titlepg_sheet, xlsx_disclaimer_sheet]}
        #===============END WORKAROUND PORTION================
        
        # params that are derived or imported from ppa_input_params.py
        self.sheets_all_rpts = sheets_all_reports_workarnd[ptype]  # params.sheets_all_reports[ptype]
        arcpy.AddMessage("This is a {} project".format(ptype))
        arcpy.AddMessage("Sheets that always go in report for this proj type: {}" \
                         .format(self.sheets_all_rpts))
        arcpy.AddMessage("Dict of sheets in all reports of this proj type, from params py file: {}" \
                         .format(params.sheets_all_reports[ptype]))
        
        self.mapimg_configs_csv = params.mapimg_configs_csv
        self.img_format = params.map_img_format # jpg, png, etc.
        self.map_placement_csv = params.map_placement_csv
        self.aprx_path = params.aprx_path
        self.proj_line_template_fc = os.path.join(params.fgdb, params.proj_line_template_fc)

        # other pre-defined class vars to use
        self.time_sufx = str(dt.datetime.now().strftime('%m%d%Y%H%M'))
        self.out_folder = arcpy.env.scratchFolder
        
        #xlsx related params
        self.xl_out_path = os.path.join(self.out_folder, self.xl_out)
        shutil.copyfile(xl_template, self.xl_out_path)
        self.xl_workbook = openpyxl.load_workbook(self.xl_out_path) #work off of a copy of the template, so template remains free. Important for multi-user reliability.


    def overwrite_df_to_xlsx(self, unused=0, start_row=0, start_col=0):  # why does there need to be an argument?
        '''Writes pandas dataframe <in_df_ to <tab_name> sheet of <xlsx_template> excel workbook.'''
        in_df = self.in_df.reset_index()
        df_records = in_df.to_records(index=False)
        
        # get header row for output
        out_header_list = [list(in_df.columns)]  # get header row for output
        
        out_data_list = [list(i) for i in df_records]  # get output data rows
    
        comb_out_list = out_header_list + out_data_list
    
        ws = self.xl_workbook[self.import_tab]
        for i, row in enumerate(comb_out_list):
            for j, val in enumerate(row):
                cell = ws.cell(row=(start_row + (i + 1)), column=(start_col + (j + 1)))
                if (cell):
                    cell.value = val
                    
    def build_configs(self):
        in_csv = self.mapimg_configs_csv
        p_map = "MapName" # map that layout and image are derived from
        p_layout = "MapLayout" # layout that will be made into image
        p_where = "SQL" # background data layer (e.g. collision heat layer)
        p_projline = "ProjLineLayer"
        
        out_config_list = []
        
        with open(in_csv, 'r') as f_in:
            reader = csv.DictReader(f_in)
            for row in reader:
                v_map = row[p_map]
                v_layout = row[p_layout]
                v_projline = row[p_projline]
                v_where = row[p_where]
                
                out_config_row = [v_map, v_layout, v_projline, v_where]
                out_config_list.append(out_config_row)
        
        return out_config_list
        

    class PrintConfig(object):
        '''each PrintConfig object has attributes: map frame, layer name, where clause'''
        def __init__(self, l_print_config, imgtyp):
            self.MapFrame = l_print_config[0]   # map/mapframe name
            self.Layout = l_print_config[1]   # layout name
            n_elements = len(l_print_config)
            if(n_elements>1):
                self.Layer = l_print_config[2]    #..layerName used to for zoomto (control ext)
            else:
                self.Layer = ""
            if(n_elements>2):
                self.Where = l_print_config[3]    #..where to get features in the layer.
            else:
                self.Where = ""
    
            self.OutputImageName = "{}.{}".format(self.MapFrame, imgtyp)
            
    def expandExtent2D(self, ext, ratio):
        '''Adjust zoom extent for map of project segment
        ext = input extent object
        ratio = how you want to change extent. Ratio > 1 zooms away from project line; <1 zooms in to project line
        '''
        try:
            # spref = ext.spatialReference
            xmin = ext.XMin
            xmax = ext.XMax
            ymin = ext.YMin
            ymax = ext.YMax 
            width = ext.width
            height = ext.height
            dx = (ratio-1.0)*width/2.0 # divided by two so that diff is split evenly between opposite sides, so featur is still center of the extent
            dy = (ratio-1.0)*height/2.0
            xxmin = xmin - dx 
            xxmax = xmax + dx
            yymin = ymin - dy 
            yymax = ymax + dy
            new_ext = arcpy.Extent(xxmin, yymin, xxmax, yymax)
        except:
            new_ext = None 
        return new_ext 
    
    # generates image files from maps
    def exportMap(self):
        arcpy.AddMessage('Generating maps for report...')
        arcpy.env.overwriteOutput = True
        try:
            # create temporary copy of APRX to not have conflicts if 2+ runs done at same time.
            aprx_temp_path = os.path.join(self.out_folder, "TEMP{}.aprx".format(int(time.perf_counter()) + 1)) 
            aprx_template_obj = arcpy.mp.ArcGISProject(self.aprx_path)
            aprx_template_obj.saveACopy(aprx_temp_path)
            
            #then manipulate the temporary copy of the APRX
            aprx = arcpy.mp.ArcGISProject(aprx_temp_path)
            
            l_print_configs = self.build_configs() # each config list for each image is [map frame name, layout frame name, project line layer name, project feature where clause]
            
            o_print_configs = []
            
            for l_print_config in l_print_configs:
                o_print_config = self.PrintConfig(l_print_config, self.img_format) #converts list vals into attributes of PrintConfig object ('o')
                o_print_configs.append(o_print_config)
            

            #insert process to overwrite display layer and append to master. This will update in all layouts using the display layer
            arcpy.DeleteFeatures_management(self.proj_line_template_fc) # delete whatever features were in the display layer
            arcpy.Append_management([self.project_fc], self.proj_line_template_fc, "NO_TEST") # then replace those features with those from user-drawn line

            for print_config in o_print_configs:
                #only thing needed for this loop is to activate each layout and pan to the desired extent and make image of it.
                layouts_aprx = [l.name for l in aprx.listLayouts()] # makes sure there's a corresponding layout in the APRX file to the layout in the CSV
                if print_config.Layout in layouts_aprx:
                    try:
                        lyt = aprx.listLayouts(print_config.Layout)[0]
                        map = aprx.listMaps(print_config.MapFrame)[0]
                        
                        if print_config.Layer != "":  # if there's a feat class for project line
                            
                            try:
                                lyr = map.listLayers(print_config.Layer)[0] # return layer object--based on layer name, not FC path
                                fl = "fl{}".format(int(time.perf_counter()))
                                if arcpy.Exists(fl):
                                    try:
                                        arcpy.Delete_management(fl)
                                    except:
                                        pass 
                                arcpy.MakeFeatureLayer_management(lyr, fl, where_clause=print_config.Where)  # make feature layer of project line
                                ext = ""
                                with arcpy.da.SearchCursor(fl, ["Shape@"]) as rows:
                                    for row in rows:
                                        geom = row[0]
                                        ext = geom.extent
                                        
                                        ext_ratio = 1.33
                                        ext_zoom = self.expandExtent2D(ext, ext_ratio)
                                        break
                                if ext_zoom != "":  # zoom to project line feature
                                    mf = lyt.listElements('MAPFRAME_ELEMENT')[0]
                                    mf.camera.setExtent(ext_zoom)
                                    mf.panToExtent(ext_zoom)

                            except:
                                msg = "{}, {}".format(arcpy.GetMessages(2), trace())
                                arcpy.AddMessage(msg)
                                
                        out_file = os.path.join(self.out_folder, print_config.OutputImageName)
                        
                        
                        if(os.path.exists(out_file)):
                            try:
                                os.remove(out_file)
                            except:
                                pass 
                        if self.img_format.lower() == 'png':
                            lyt.exportToPNG(out_file)
                        elif self.img_format.lower() == 'jpg':
                            lyt.exportToJPEG(out_file) # after zooming in, export the layout to a JPG
                        else:
                            arcpy.AddWarning("Map image {} not created. Must be PNG or JPG.".format(out_file))
                    except:
                        msg = "{}, {}".format(arcpy.GetMessages(2), trace())
                        arcpy.AddMessage(msg)
                        print(msg)
                else:
                    continue # if specified layout isn't in APRX project file, skip to next map
            t_returns = (params.msg_ok,)
        except:
            msg = "{}, {}".format(arcpy.GetMessages(2), trace())
            print(msg)
            arcpy.AddWarning(msg)
            t_returns = (msg,)
    
    def insert_image_xlsx(self, wb, sheet_name, rownum, col_letter, img_file):
        '''inserts image into specified sheet and cell within Excel workbook'''
        ws = wb[sheet_name]
        cell = '{}{}'.format(col_letter, rownum) # will be where upper left corner of image placed
        img_obj = Image(img_file)
        ws.add_image(img_obj, cell)
        
    def move_sheets(self, sheets_to_move):
    
        sheet_objs = self.xl_workbook._sheets
        # title_posn = wb.sheetnames.index(params.xlsx_titlepg_sheet)
        disclaimer_posn = self.xl_workbook.sheetnames.index(params.xlsx_disclaimer_sheet)
        
        for i, sheet in enumerate(sheets_to_move):
            start_pos = self.xl_workbook.sheetnames.index(sheet)
            posns_to_move = i + 1
            destination = disclaimer_posn + posns_to_move
            
            sheet_obj2move = sheet_objs.pop(start_pos) # cut sheet out of original position
            sheet_objs.insert(destination, sheet_obj2move) # paste into desired position

    def color_sheets(self, sheet_names):
        for sheet_name in sheet_names:
            sheet_posn = self.xl_workbook.sheetnames.index(sheet_name)
            sheet_obj = self.xl_workbook._sheets[sheet_posn]
            sheet_obj.sheet_properties.tabColor = "45b045" # RGB color code
        
    def make_new_excel(self):
        try:
            '''takes excel template > writes new values to import/updates charts, then inserts indicated images at specified locations'''
            self.overwrite_df_to_xlsx(self) # write data to import tab
            
            self.exportMap() # generates all needed maps as images in scratch folder
            
            if self.map_placement_csv:
                # indicate which sheets and cells to insert maps in, based on config CSV and template XLSX (not output XLSX)
                # add conditional or try/catch series to skip rows that are missing info (e.g. cell location, img file name, etc.)
                bookname = os.path.basename(self.xl_template)
                mapkey = pd.read_csv(self.map_placement_csv)
                col_mapfile = 'MapImgFile'
                col_rownum = 'RowNum'
                col_colletter = 'ColNum'
                col_sheet = 'Tab'
                
                 # filter master config table to only get tabs for workbook corresponding to specified project type
                # and where there are non-null values. Null values will trigger an error.
                mapkey_dict_list = mapkey.loc[(mapkey['Report'] == bookname) \
                                              & (pd.notnull(mapkey[col_mapfile])) \
                                            & (pd.notnull(mapkey[col_rownum])) \
                                            & (pd.notnull(mapkey[col_colletter]))] \
                    .to_dict(orient='records')
                
                for i in mapkey_dict_list:
                    imgfile = "{}.{}".format(i[col_mapfile], self.img_format)
                    sheet = i[col_sheet]
                    row = int(i[col_rownum]) # needs to be int value, not float, for openpyxl to use as cell reference
                    col = i[col_colletter]
                            
                    # row is valid only if it has all necessary image attributes (not null values)
                    valid_row = []  # [list of boolean True/False values. True means the is_valid condition is met]
                    for i in [imgfile, sheet, row, col]:
                        is_valid = isinstance(i, str) | (i != math.nan)
                        valid_row.append(is_valid)
                    
                    if pd.Series(valid_row).product() == 1: # run the insert_image function only if valid config values are provided in CSV (i.e., they meet the is_valid condition above)
                        imgfilepath = os.path.join(self.out_folder, imgfile)
                        self.insert_image_xlsx(self.xl_workbook, sheet, row, col, imgfilepath)
                    else:
                        continue # skip to next image if the row has no image specified (i.e., if it's a placeholder row only)

            # write report generation time stamp to cell on title page
            tstamp = dt.datetime.strftime(dt.datetime.now(),"%m-%d-%Y %H:%M")
            self.xl_workbook[params.xlsx_titlepg_sheet][params.tstamp_cell] = tstamp
            
            # move sheets with user-selected perf outcomes to be at front Final order will be: title, disclaimer, <3 perf sheets>, sometimes soc equity outcome
            
            selected_report_sheets = self.selecd_po_sheets
            
            # social equity sheet is not in all outputs. if it is, add it.
            if params.xlsx_socequity_sheet in self.sheets_all_rpts:
                selected_report_sheets.append(params.xlsx_socequity_sheet)
                
            self.move_sheets(selected_report_sheets)
            
            # then highlight all the tabs the user should print out (title, selected perf outcomes, etc)
            # leader_sheets = [params.xlsx_titlepg_sheet, params.xlsx_disclaimer_sheet]
            # sheets_to_color = leader_sheets + selected_report_sheets  # [title, disclaimer, <3 user-selected outcome reports>, social equity outcome report]

            sheets_to_color = self.sheets_all_rpts  # [title, disclaimer, potentially soc equity sheet]

            # insert user-selected perf outcome sheets to get [title, disclaimer, <3 perf outcomes>, potentially soc equity sheet]
            for i, sheet in enumerate(self.selecd_po_sheets):
                idxstart = sheets_to_color.index(params.xlsx_disclaimer_sheet)
                ins_posn = idxstart + i + 1
                sheets_to_color.insert(ins_posn, sheet)
                
            self.color_sheets(sheets_to_color)
            
            self.xl_workbook.save(self.xl_out_path)
            
            t_returns = (params.msg_ok, self.xl_out_path)
        except:
            msg = "{}".format(trace())
            t_returns = (params.msg_fail, msg)   
        finally:
            if self.xl_workbook.close() != None:
                self.xl_workbook.close()

            del self.sheets_all_rpts 
            gc.collect()
        
        return t_returns
        
    def make_pdf(self):
        wb = None
        out_pdf_final_name = "Rpt_{}{}.pdf".format(self.proj_name, self.time_sufx)
        out_pdf_final = os.path.join(self.out_folder, out_pdf_final_name)
            
        if os.path.exists(out_pdf_final):
            try:
                os.remove(out_pdf_final)
            except:
                out_pdf_final_name = "Rpt{}{}v2.pdf".format(self.proj_name, self.time_sufx)
                out_pdf_final = os.path.join(self.out_folder, out_pdf_final_name)
            
        # make new excel file if it doesn't already exist
        if not os.path.exists(self.xl_out):
            self.make_new_excel()        
            
        try:            
            arcpy.AddMessage("Publishing to PDF...")
            
            # make excel workbook with project outputs
            xw.App.visible = True # must be set to True or else it won't work            
            wb = xw.Book(self.xl_out_path)
            
            # make single list of pages that must be in all reports + pages that user selected (selecd_po_sheets)
            out_sheets = self.sheets_all_rpts + self.selecd_po_sheets
                
            l_out_pdfs = [] # will be list of output PDF file names, PDFs in this list will be combined. Need list step for sorting by sheet name A-Z
            
            pdf_final = arcpy.mp.PDFDocumentCreate(out_pdf_final) # instantiate arcpy PDFDocumentCreate object
                
            # write user-specified sheets to PDFs
            for s in out_sheets:
                out_sheet = wb.sheets[s]
                pdf_out = os.path.join(self.out_folder, 'Sheet_{}_{}.pdf'.format(s, self.time_sufx))
                out_sheet.api.ExportAsFixedFormat(0, pdf_out)
                l_out_pdfs.append(pdf_out)
            
            l_out_pdfs = sorted(l_out_pdfs) # sort by sheet name so that PDFs append together in correct order.
            
            for singlesheet_pdf in l_out_pdfs:
                pdf_final.appendPages(singlesheet_pdf) # append to master PDF object
                try:
                    os.remove(singlesheet_pdf)  # not necessary for online version because scratch folder is temporary
                except:
                    pass
                
            pdf_final.saveAndClose()
            
            t_returns = (params.msg_ok, self.xl_out_path, out_pdf_final) # if successful, return output excel and output PDF
        except:
            msg = "{}".format(trace())
            arcpy.AddMessage(msg)
            
            t_returns = (params.msg_fail, msg, self.xl_out_path) # if fail, return error message that PDF didn't make it, but that Excel still made it
        finally: # always runs, even if 'try' runs successfully.
            if wb != None:  # only closes wb object if it was instantiated.
                wb.close()
            gc.collect()
            
        return t_returns





