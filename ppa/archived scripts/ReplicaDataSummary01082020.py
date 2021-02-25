# -*- coding: utf-8 -*-
"""
Created on Wed Dec 11 13:46:21 2019

@author: dconly
"""
import os
import datetime as dt

import pandas as pd
from openpyxl import load_workbook
import arcpy

import ppa_input_params as p
import PPA2_masterBYFY_tripshed as tripshed


def overwrite_df_to_xlsx(in_df, xlsx_template, xlsx_out, tab_name, start_row=0, start_col=0):
    
    df_records = in_df.to_records()
    out_header_list = [list(in_df.columns)]  # get header row for output
    out_data_list = [list(i) for i in df_records]  # get output data rows
    
    comb_out_list = out_header_list + out_data_list
    
    wb = load_workbook(xlsx_template)
    ws = wb[tab_name]
    for i, row in enumerate(comb_out_list):
        if i > 0:
            for j, val in enumerate(row):
                cell = ws.cell(row = (start_row + (i + 1)), column = (start_col + (j + 1)))
                if(cell):
                    cell.value = val
    wb.save(xlsx_out)
    

def make_fl_conditional(fc, fl):
    if arcpy.Exists(fl):
        arcpy.Delete_management(fl)
    arcpy.MakeFeatureLayer_management(fc, fl)
    

def esri_field_exists(in_tbl, field_name):
    fields = [f.name for f in arcpy.ListFields(in_tbl)]
    if field_name in fields:
        return True
    else:
        return False

def make_tripdata_df(in_files):
    # if single CSV, read in to pandas df; if multiple CSVs, read sequentially and append together into single df
    if len(in_files) == 1:
        out_df = pd.read_csv(in_files[0])
    else:
        out_df = pd.read_csv(in_files[0])
        for file in in_files[1:]:
            out_df2 = pd.read_csv(file)
            out_df = out_df.append(out_df2)
    
    return out_df

def summarize_tripdf(in_df, groupby_field, val_field, agg_fxn):
    # get totals for all trips (e.g. total trips x mode), irrespective of geography/filtering. 
    df_gb = in_df.groupby(groupby_field)[val_field].agg(agg_fxn)
    df_gb = pd.DataFrame(df_gb).rename(columns={'{}'.format(val_field): '{}'.format(agg_fxn)})
    df_gb['category'] = groupby_field
    
    df_gb['pct'] = df_gb[agg_fxn] / df_gb[agg_fxn].sum()
    
    return df_gb

def get_poly_data(in_df, val_field, agg_fxn, groupby_field, categ_cols=[], case_field=None):
    '''make dataframe summarizing items by desired polygon ID field (e.g. trips by block group ID)
    The categ_cols argument specifies what you want the columns to be (e.g., trips by mode, by purpose, etc.)
    '''
    
    col_tottrips = 'tot_trips'
    col_pcttrips = 'pct_of_trips'
    col_trippctlrank = 'trips_pctlrank'
    
    piv = in_df.pivot_table(values=val_field, index=groupby_field, columns=case_field, 
                            aggfunc=agg_fxn)
    piv = piv.reset_index()
    
    # option to break out trips in each poly by some category (e.g., mode or purpose)
    if len(categ_cols) > 0:
        tblmodes = [col for col in categ_cols if col in piv.columns]
        
        piv[col_tottrips] = piv[tblmodes].sum(axis=1)  # count of trips in each poly
    else:
        piv = piv.rename(columns={val_field: col_tottrips}) # or just give total trips from each poly
    
    piv[col_pcttrips] = piv[col_tottrips] / piv[col_tottrips].sum()  # pct of total trips from each poly
    
    # percentile rank of each poly in terms of how many trips it produces
    # e.g., 100th percentile means the poly makes more trips than any of the other polys
    piv[col_trippctlrank] = piv[col_tottrips].rank(method='min', pct=True)

    return piv


# filter input polygon set to only retrieve polygons that capture some majority share (cutoff share, e.g., 90%)of the total trips
# but selected in descending order of how many trips were created.
def filter_cumulpct(in_df, sort_col, cutoff):
    df_sorted = in_df.sort_values(sort_col, ascending=False)
    
    col_cumsumpct = 'cumul_sum'
    df_sorted[col_cumsumpct] = df_sorted[sort_col].cumsum()
    
    df_out = df_sorted[df_sorted[col_cumsumpct] <= cutoff]
    
    return df_out


def create_tripshed_poly(in_poly_fc, out_poly_fc, poly_id_field, in_df, df_grouby_field):
    
    # convert numpy (pandas) datatypes to ESRI data types {numpy type: ESRI type}
    dtype_conv_dict = {'float64': 'FLOAT', 'object': 'TEXT', 'int64': 'LONG', 
                       'String': 'TEXT', 'OID': 'LONG', 'Single': 'DOUBLE', 
                       'Integer': 'LONG'}
    
    #make copy of base input poly fc that only has features whose IDs are in the dataframe
    fl_input_polys = 'fl_input_polys'
    make_fl_conditional(in_poly_fc, fl_input_polys)
    
    df_ids = tuple(in_df[df_grouby_field])
    
    sql = "{} IN {}".format(poly_id_field, df_ids)
    arcpy.SelectLayerByAttribute_management(fl_input_polys, "NEW_SELECTION", sql)
    
    arcpy.CopyFeatures_management(fl_input_polys, out_poly_fc)

    # add dataframe fields to the trip shed polygon set
            
    #dict of {field: field data type} for input dataframe
    fields_dtype_dict = {col:str(in_df[col].dtype) for col in in_df.columns}
    
    # populate those fields with the dataframe data
    for field in fields_dtype_dict.keys():
        print("adding {} column and data...".format(field))
        field_vals = list(in_df[field]) # get list of values for desired field
        fld_dict = dict(zip(df_ids, field_vals))
        
        fdtype_numpy = fields_dtype_dict[field]
        fdtype_esri = dtype_conv_dict[fdtype_numpy]
        
        # add a field, if needed, to the polygon feature class for the values being added
        if esri_field_exists(out_poly_fc, field):
            pass
        else:
            arcpy.AddField_management(out_poly_fc, field, fdtype_esri)
            
        # populate the field with the appropriate values
        with arcpy.da.UpdateCursor(out_poly_fc, [poly_id_field, field]) as cur:
            for row in cur:
                join_id = row[0]
                if fld_dict.get(join_id) is None:
                    pass
                else:
                    row[1] = fld_dict[join_id]
                    cur.updateRow(row)


# get PPA buffer data for trip shed
def make_trip_shed_report(in_tripdata_files, tripdata_val_field, tripdata_agg_fxn, tripdata_groupby_field,
                   in_poly_fc, out_poly_fc, poly_id_field, analysis_years, tripdata_case_fields):

    df_col_trip_pct = 'pct_of_trips'
    
    # import CSV trip data into dataframe
    df_tripdata = make_tripdata_df(in_tripdata_files)
    
    # get splits of trips by mode and trips by purpose
    df_linktripsummary = summarize_tripdf(df_tripdata, tripdata_case_fields[0], tripdata_val_field, tripdata_agg_fxn)
    
    for f in tripdata_case_fields[1:]:
        df_linktripsummary = df_linktripsummary.append(summarize_tripdf(df_tripdata, f, tripdata_val_field, tripdata_agg_fxn))
    
    # summarize by polygon and whatever case value
    "aggregating Replica trip data for each polygon within trip shed..."
    df_groupd_tripdata = get_poly_data(df_tripdata, tripdata_val_field, tripdata_agg_fxn, 
                                       tripdata_groupby_field, [])

    # filter out block groups that don't meet inclusion criteria (remove polys that have few trips, but keep enough polys to capture X percent of all trips)
    df_outdata = filter_cumulpct(df_groupd_tripdata, df_col_trip_pct, 0.8)


    # make new polygon feature class with trip data
    create_tripshed_poly(in_poly_fc, out_poly_fc, poly_id_field, df_outdata, tripdata_groupby_field)
    print("created polygon feature class {}.".format(out_poly_fc))
    
    # get PPA buffer data for trip shed
    print("summarizing PPA metrics for trip shed polygon...")
    df_tsheddata = tripshed.get_tripshed_data(out_poly_fc, p.ptype_area_agg, analysis_years, p.aggvals_csv, base_dict={})

    return df_tsheddata, df_linktripsummary

if __name__ == '__main__':
    
    # ------------------USER INPUTS----------------------------------------
    
    arcpy.env.workspace = r'I:\Projects\Darren\PPA_V2_GIS\PPA_V2.gdb'
    dir_tripdata = r'Q:\ProjectLevelPerformanceAssessment\PPAv2\Replica\ReplicaDownloads'
    
    #community-type and region-level values for comparison to project-level values
    aggvals_csv = r"Q:\ProjectLevelPerformanceAssessment\PPAv2\PPA2_0_code\PPA2\AggValCSVs\Agg_ppa_vals01022020_1004.csv"
    
    proj_name = input('Enter project name (numbers, letters, and underscores only): ')
    
    tripdata_files = ['yolo_80causeway_thu_0000_1159.zip',
                      'yolo_80causeway_thu_1200_2359.zip']
    
    csvcol_bgid = 'origin_blockgroup_id'  # Replica/big data block group ID column
    csvcol_mode = 'trip_primary_mode'  # Replica/big data trip mode column
    csvcol_purpose = 'travel_purpose'  # Replica/big data trip purpose column
    
    csvcol_valfield = 'trip_start_time' # field for which you want to aggregate values
    val_aggn_type = 'count'  # how you want to aggregate the values field (e.g. count of values, sum of values, avg, etc.)
    
    # feature class of polygons to which you'll join data to based on ID field
    fc_bg_in = "BlockGroups2010"
    fc_poly_id_field = "GEOID10"
    
    years = [2016, 2040]
    
    xlsx_template = r"Q:\ProjectLevelPerformanceAssessment\PPAv2\Replica\Replica_Summary_Template.xlsx"
    xlsx_out_dir = r"Q:\ProjectLevelPerformanceAssessment\PPAv2\Replica"

    #------------RUN SCRIPT------------------------------------
    timesufx = str(dt.datetime.now().strftime('%m%d%Y_%H%M'))
    os.chdir(dir_tripdata)
    
    xlsx_out = '{}_TripShedAnalysis_{}.xlsx'.format(proj_name, timesufx)
    xlsx_out = os.path.join(xlsx_out_dir, xlsx_out)

    fc_tripshed_out = "TripShed_{}{}".format(proj_name, timesufx)
    
    df_tshed_data, link_trip_summary = make_trip_shed_report(tripdata_files, csvcol_valfield, val_aggn_type, csvcol_bgid,
                   fc_bg_in, fc_tripshed_out, fc_poly_id_field, years, tripdata_case_fields=[csvcol_mode, csvcol_purpose])
    
    output_csv = r'Q:\ProjectLevelPerformanceAssessment\PPAv2\PPA2_0_code\PPA2\ProjectValCSVs\PPA_TripShed_{}_{}.csv'.format(
        proj_name, timesufx)
    
    df_tshed_data.to_csv(output_csv)
    print("Success! Output file is {}".format(output_csv))
    
    dfs_tabs_dict = {'df_tshed_data': df_tshed_data, 'link_trip_summary': link_trip_summary}
    
    overwrite_df_to_xlsx(df_tshed_data, xlsx_template, xlsx_out, 'df_tshed_data', start_row=0, start_col=0)
    overwrite_df_to_xlsx(link_trip_summary, xlsx_out, xlsx_out, 'link_trip_summary', start_row=0, start_col=0)
    
    
    