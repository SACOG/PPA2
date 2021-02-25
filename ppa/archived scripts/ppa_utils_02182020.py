# --------------------------------
# Name: utils.py
# Purpose: Provides general PPA functions that are used throughout various PPA scripts and are not specific to any one PPA script
#
#
# Author: Darren Conly
# Last Updated: <date>
# Updated by: <name>
# Copyright:   (c) SACOG
# Python Version: 3.x
# --------------------------------

import os
import pdb
import sys
import datetime as dt
import time
import gc
import csv
import math

import openpyxl
from openpyxl.drawing.image import Image
import xlwings as xw
import pandas as pd
import arcpy

import ppa_input_params as params


def trace():
    import traceback, inspect
    tb = sys.exc_info()[2]
    tbinfo = traceback.format_tb(tb)[0]
    # script name + line number
    line = tbinfo.split(", ")[1]
    filename = inspect.getfile(inspect.currentframe())
    # Get Python syntax error
    synerror = traceback.format_exc().splitlines()[-1]
    return line, filename, synerror


def make_fl_conditional(fc, fl):
    '''check if a feature layer name exists; if it does, delete feature layer and remake it.
    purpose is to ensure the feature layer name corresponds to the correct feature class.'''
    if arcpy.Exists(fl):
        arcpy.Delete_management(fl)
    arcpy.MakeFeatureLayer_management(fc, fl)
    
    
def esri_field_exists(in_tbl, field_name):
    fields = [f.name for f in arcpy.ListFields(in_tbl)]
    if field_name in fields:
        return True
    else:
        return False


def esri_object_to_df(in_esri_obj, esri_obj_fields, index_field=None):
    '''converts esri gdb table, feature class, feature layer, or SHP to pandas dataframe'''
    data_rows = []
    with arcpy.da.SearchCursor(in_esri_obj, esri_obj_fields) as cur:
        for row in cur:
            out_row = list(row)
            data_rows.append(out_row)

    out_df = pd.DataFrame(data_rows, index=index_field, columns=esri_obj_fields)
    return out_df


def return_perf_outcomes_options(project_type):
    arcpy.AddMessage(project_type)
    xlsx = params.type_template_dict[project_type]
    xlsx_path = os.path.join(params.template_dir, xlsx)
    
    wb = openpyxl.load_workbook(xlsx_path)
    sheets = wb.sheetnames
    
    # values in this list will be the potential performance outcomes from which users can choose
    perf_outcomes = [s for s in sheets if s not in params.sheets_all_reports] 
    return perf_outcomes
    
    
def rename_dict_keys(dict_in, new_key_dict):
    '''if dict in = {0:1} and dict out supposed to be {'zero':1}, this function renames the key accordingly per
    the new_key_dict (which for this example would be {0:'zero'}'''
    dict_out = {}
    for k, v in new_key_dict.items():
        if k in list(dict_in.keys()):
            dict_out[v] = dict_in[k]
        else:
            dict_out[v] = 0
    return dict_out



def join_xl_import_template(template_xlsx, template_sheet, in_df):
    '''takes in import tab of destination Excel sheet, then left joins to desired output dataframe to ensure that
    output CSV has same rows every time, even if data frame that you're joining doesn't
    have all records'''
    df_template = pd.read_excel(template_xlsx, template_sheet)
    df_template = pd.DataFrame(df_template[df_template.columns[0]]) # get rid of all columns except for data items column
    df_template = df_template.set_index(df_template.columns[0]) # set data items column to be the index
    
    df_out = df_template.join(in_df) # left join the df from import sheet template to the df with data based on index values
    
    return df_out


class Publish(object):
    def __init__(self, in_df, xl_template, import_tab, xl_out, project_fc, ptype, xlsheets_to_pdf=None, 
                 proj_name='UnnamedProject'):
        # params from input arguments
        self.in_df = in_df
        self.xl_template = xl_template
        self.import_tab = import_tab
        self.xl_out = xl_out
        self.xlsheets_to_pdf = xlsheets_to_pdf
        self.proj_name = proj_name
        self.project_fc = project_fc
        
        # params that are derived or imported from ppa_input_params.py
        self.xl_template_workbook = openpyxl.load_workbook(self.xl_template)
        self.time_sufx = str(dt.datetime.now().strftime('%m%d%Y%H%M'))
        self.sheets_all_rpts = params.sheets_all_reports[ptype]
        self.out_folder = arcpy.env.scratchFolder
        self.xl_out_path = os.path.join(self.out_folder, self.xl_out)
        self.mapimg_configs_csv = params.mapimg_configs_csv
        self.map_placement_csv = params.map_placement_csv
        self.aprx_path = params.aprx_path


    def overwrite_df_to_xlsx(self, unused=0, start_row=0, start_col=0):  # why does there need to be an argument?
        '''Writes pandas dataframe <in_df_ to <tab_name> sheet of <xlsx_template> excel workbook.'''
        in_df = self.in_df.reset_index()
        df_records = in_df.to_records(index=False)
        
        # get header row for output
        out_header_list = [list(in_df.columns)]  # get header row for output
        
        out_data_list = [list(i) for i in df_records]  # get output data rows
    
        comb_out_list = out_header_list + out_data_list
    
        ws = self.xl_template_workbook[self.import_tab]
        for i, row in enumerate(comb_out_list):
            for j, val in enumerate(row):
                cell = ws.cell(row=(start_row + (i + 1)), column=(start_col + (j + 1)))
                if (cell):
                    cell.value = val
                    
    def build_configs(self):
        in_csv = self.mapimg_configs_csv
        p_map = "MapName" # map that layout and image are derived from
        p_layout = "MapLayout" # layout that will be made into image
        p_where = "SQL" # background data layer (e.g. collision heat layer)
        
        out_config_list = []
        
        with open(in_csv, 'r') as f_in:
            reader = csv.DictReader(f_in)
            for row in reader:
                v_map = row[p_map]
                v_layout = row[p_layout]
                v_where = row[p_where]
                out_config_row = [v_map, v_layout, v_where]
                out_config_list.append(out_config_row)
        
        return out_config_list
        

    class PrintConfig(object):
        '''each PrintConfig object has attributes: map frame, layer name, where clause'''
        def __init__(self, l_print_config, project_layer, imgtyp):
            self.MapFrame = l_print_config[0]   # map/mapframe name
            self.Layout = l_print_config[1]   # layout name
            n_elements = len(l_print_config)
            if(n_elements>1):
                self.Layer = project_layer    #..layerName used to for zoomto (control ext)
            else:
                self.Layer = ""
            if(n_elements>2):
                self.Where = l_print_config[2]    #..where to get features in the layer.
            else:
                self.Where = ""
    
            self.OutputImageName = "{}.{}".format(self.MapFrame, imgtyp)
    
    # generates image files from maps
    def exportmap(self):
        arcpy.AddMessage('Generating maps for report...')
        image_format = "jpg"
        arcpy.env.overwriteOutput = True
        out_images = ""
        try:
            aprx = arcpy.mp.ArcGISProject(self.aprx_path)
            l_print_configs = self.build_configs() # each config list for each image is [map frame name, layout frame name, project line layer name, project feature where clause]
            
            o_print_configs = []
            
            project = self.project_fc #assigning to variable so no confusion over which class it project_fc belongs to
            for l_print_config in l_print_configs:
                o_print_config = self.PrintConfig(l_print_config, project, image_format) #converts list vals into attributes of PrintConfig object ('o')
                o_print_configs.append(o_print_config)
            
            pdb.set_trace()
            for print_config in o_print_configs:
                layouts_aprx = [l.name for l in aprx.listLayouts()] # makes sure there's a corresponding layout in the APRX file to the layout in the CSV
                if print_config.Layout in layouts_aprx:
                    try:
                        lyt = aprx.listLayouts(print_config.Layout)[0]
                        map = aprx.listMaps(print_config.MapFrame)[0]
                        if print_config.Layer != "":  # if there's a layer
                            try:
                                # somewhere in here, need to insert the layer representing the project line
                                # https://pro.arcgis.com/en/pro-app/arcpy/mapping/map-class.htm
                                map_layers = [l.dataSource for l in map.listLayers()]
                                if print_config.Layer in map_layers:
                                    map.removeLayer(print_config.Layer)
                                map.addDataFromPath(print_config.Layer)
                                arcpy.AddMessage("Added layer {} to map {}".format(print_config.Layer, print_config.MapFrame))
                                print("Added layer {} to map {}".format(print_config.Layer, print_config.MapFrame))
                                
                                lyr = map.listLayers(print_config.Layer)[0] # return layer object--based on layer name, not FC path
                                fl = "fl{}".format(int(time.clock()))
                                if arcpy.Exists(fl):
                                    try:
                                        arcpy.Delete_management(fl)
                                    except:
                                        pass 
                                arcpy.MakeFeatureLayer_management(lyr, fl, where_clause=print_config.Where)  # make feature layer of project line
                                arcpy.AddMessage("{} {}".format(arcpy.GetCount_management(fl)[0], print_config.Where))
                                ext = ""
                                with arcpy.da.SearchCursor(fl, ["Shape@"]) as rows:
                                    for row in rows:
                                        geom = row[0]
                                        ext = geom.extent
                                        break
                                if ext != "":  # zoom to project line feature
                                    mf = lyt.listElements('MAPFRAME_ELEMENT')[0]
                                    mf.camera.setExtent(ext)
                                    mf.panToExtent(ext)
                            except:
                                msg = "{}, {}".format(arcpy.GetMessages(2), trace())
                                arcpy.AddMessage(msg)
                                
                        out_file = os.path.join(self.out_folder, print_config.OutputImageName)
                        
                        if(os.path.exists(out_file)):
                            try:
                                os.remove(out_file)
                            except:
                                pass 
                        lyt.exportToJPEG(out_file) # after zooming in, export the layout to a JPG
                        
                        # make semicolon-delim'd list of output JPEGs
                        if(out_images==""):
                            out_images = out_file
                        else:
                            out_images = "{};{}".format(out_images, out_file)
                    except:
                        msg = "{}, {}".format(arcpy.GetMessages(2), trace())
                        arcpy.AddMessage(msg)
                else:
                    continue # if specified layout isn't in APRX project file, skip to next map
            t_returns = (params.msg_ok, out_images)
        except:
            msg = "{}, {}".format(arcpy.GetMessages(2), trace())
            arcpy.AddWarning(msg)
            t_returns = (msg,)
    
    def insert_image_xlsx(self, wb, sheet_name, rownum, col_letter, img_file):
        '''inserts image into specified sheet and cell within Excel workbook'''
        ws = wb[sheet_name]
        cell = '{}{}'.format(col_letter, rownum) # will be where upper left corner of image placed
        img_obj = Image(img_file)
        ws.add_image(img_obj, cell)
        
    def make_new_excel(self):
        '''takes excel template > writes new values to import/updates charts, then inserts indicated images at specified locations'''
        self.overwrite_df_to_xlsx(self) # write data to import tab
        
        self.exportmap() # generates all needed maps as images in scratch folder
        
        if self.map_placement_csv:
            # indicate which sheets and cells to insert maps in, based on config CSV and template XLSX (not output XLSX)
            # add conditional or try/catch series to skip rows that are missing info (e.g. cell location, img file name, etc.)
            bookname = os.path.basename(self.xl_template)
            mapkey = pd.read_csv(self.map_placement_csv)
            col_mapfile = 'MapImgFile'
            col_rownum = 'RowNum'
            col_colletter = 'ColNum'
            col_sheet = 'Tab'
            
             # filter master table to only get tabs for workbook corresponding to specified project type
            # and where there are non-null values 
            mapkey_dict_list = mapkey.loc[(mapkey['Report'] == bookname) \
                                          & (pd.notnull(mapkey[col_mapfile])) \
                                        & (pd.notnull(mapkey[col_rownum])) \
                                        & (pd.notnull(mapkey[col_colletter]))] \
                .to_dict(orient='records')
            
            for i in mapkey_dict_list:
                imgfile = i[col_mapfile]
                sheet = i[col_sheet]
                row = int(i[col_rownum]) #needs to be int value, not float, for openpyxl to use as cell reference
                col = i[col_colletter]
                        
                # row is valid spec row only if it has all necessary image attributes (not null values)
                valid_row = []
                for i in [imgfile, sheet, row, col]:
                    is_valid = isinstance(i, str) | (i != math.nan)
                    valid_row.append(is_valid)
                
                if pd.Series(valid_row).product() == 1:
                    imgfilepath = os.path.join(self.out_folder, imgfile)
                    self.insert_image_xlsx(self.xl_template_workbook, sheet, row, col, imgfilepath)
                else:
                    continue # skip to next image if the row has no image specified (placeholder row only)
        
        self.xl_template_workbook.save(self.xl_out_path)
        self.xl_template_workbook.close()
        
    def make_pdf(self):
        wb = None
        out_pdf_final_name = "Rpt_{}{}.pdf".format(self.proj_name, self.time_sufx)
        out_pdf_final = os.path.join(self.out_folder, out_pdf_final_name)
            
        if os.path.exists(out_pdf_final):
            try:
                os.remove(out_pdf_final)
            except:
                out_pdf_final_name = "Rpt{}{}v2.pdf".format(self.proj_name, self.time_sufx)
                out_pdf_final = os.path.join(self.out_folder, out_pdf_final_name)
            
        # make new excel file if it doesn't already exist
        if not os.path.exists(self.xl_out):
            self.make_new_excel()        
            
        try:            
            arcpy.AddMessage("Publishing to PDF...")
            
            # make excel workbook with project outputs
            xw.App.visible = True # must be set to True or else it won't work            
            wb = xw.Book(self.xl_out_path)
            
            # make single list of pages that must be in all reports + pages that user selected (xlsheets_to_pdf)
            out_sheets = self.sheets_all_rpts + self.xlsheets_to_pdf
                
            l_out_pdfs = [] # will be list of output PDF file names, PDFs in this list will be combined. Need list step for sorting by sheet name A-Z
            
            pdf_final = arcpy.mp.PDFDocumentCreate(out_pdf_final) # instantiate arcpy PDFDocumentCreate object
                
            # write user-specified sheets to PDFs
            for s in out_sheets:
                out_sheet = wb.sheets[s]
                pdf_out = os.path.join(self.out_folder, 'Sheet_{}_{}.pdf'.format(s, self.time_sufx))
                out_sheet.api.ExportAsFixedFormat(0, pdf_out)
                l_out_pdfs.append(pdf_out)
            
            l_out_pdfs = sorted(l_out_pdfs) # sort by sheet name so that PDFs append together in correct order.
            
            for singlesheet_pdf in l_out_pdfs:
                pdf_final.appendPages(singlesheet_pdf) # append to master PDF object
                try:
                    os.remove(singlesheet_pdf)  # not necessary for online version because scratch folder is temporary
                except:
                    pass
                
            pdf_final.saveAndClose()
            
            t_returns = (params.msg_ok, out_pdf_final, self.xl_out_path) # if successful, return output excel and output PDF
        except:
            msg = "{}".format(trace())
            arcpy.AddMessage(msg)
            
            t_returns = (msg, self.xl_out_path) # if fail, return error message that PDF didn't make it, but that Excel still made it
        finally: # always runs, even if 'try' runs successfully.
            if wb != None:  # only closes wb object if it was instantiated.
                wb.close()
            gc.collect()
            
        return t_returns




