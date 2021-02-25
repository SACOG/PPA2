# --------------------------------
# Name: utils.py
# Purpose: Provides general PPA functions that are used throughout various PPA scripts and are not specific to any one PPA script
#
#
# Author: Darren Conly
# Last Updated: <date>
# Updated by: <name>
# Copyright:   (c) SACOG
# Python Version: 3.x
# --------------------------------

import os
import sys
import datetime as dt
import gc

import openpyxl
from openpyxl.drawing.image import Image
import xlwings as xw
import pandas as pd
import arcpy

import ppa_input_params as params


def trace():
    import traceback, inspect
    tb = sys.exc_info()[2]
    tbinfo = traceback.format_tb(tb)[0]
    # script name + line number
    line = tbinfo.split(", ")[1]
    filename = inspect.getfile(inspect.currentframe())
    # Get Python syntax error
    synerror = traceback.format_exc().splitlines()[-1]
    return line, filename, synerror


def make_fl_conditional(fc, fl):
    '''check if a feature layer name exists; if it does, delete feature layer and remake it.
    purpose is to ensure the feature layer name corresponds to the correct feature class.'''
    if arcpy.Exists(fl):
        arcpy.Delete_management(fl)
    arcpy.MakeFeatureLayer_management(fc, fl)
    
    
def esri_field_exists(in_tbl, field_name):
    fields = [f.name for f in arcpy.ListFields(in_tbl)]
    if field_name in fields:
        return True
    else:
        return False


def esri_object_to_df(in_esri_obj, esri_obj_fields, index_field=None):
    '''converts esri gdb table, feature class, feature layer, or SHP to pandas dataframe'''
    data_rows = []
    with arcpy.da.SearchCursor(in_esri_obj, esri_obj_fields) as cur:
        for row in cur:
            out_row = list(row)
            data_rows.append(out_row)

    out_df = pd.DataFrame(data_rows, index=index_field, columns=esri_obj_fields)
    return out_df


def return_perf_outcomes_options(project_type):
    arcpy.AddMessage(project_type)
    xlsx = params.type_template_dict[project_type]
    xlsx_path = os.path.join(params.template_dir, xlsx)
    
    wb = openpyxl.load_workbook(xlsx_path)
    sheets = wb.sheetnames
    
    # values in this list will be the potential performance outcomes from which users can choose
    perf_outcomes = [s for s in sheets if s not in params.sheets_all_reports] 
    return perf_outcomes
    
    
def rename_dict_keys(dict_in, new_key_dict):
    '''if dict in = {0:1} and dict out supposed to be {'zero':1}, this function renames the key accordingly per
    the new_key_dict (which for this example would be {0:'zero'}'''
    dict_out = {}
    for k, v in new_key_dict.items():
        if k in list(dict_in.keys()):
            dict_out[v] = dict_in[k]
        else:
            dict_out[v] = 0
    return dict_out



def join_xl_import_template(template_xlsx, template_sheet, in_df):
    '''takes in import tab of destination Excel sheet, then left joins to desired output dataframe to ensure that
    output CSV has same rows every time, even if data frame that you're joining doesn't
    have all records'''
    df_template = pd.read_excel(template_xlsx, template_sheet)
    df_template = pd.DataFrame(df_template[df_template.columns[0]]) # get rid of all columns except for data items column
    df_template = df_template.set_index(df_template.columns[0]) # set data items column to be the index
    
    df_out = df_template.join(in_df) # left join the df from import sheet template to the df with data based on index values
    
    return df_out


class Publish(object):
    def __init__(self, in_df, xl_template, import_tab, xl_out, xlsheets_to_pdf=None, map_key_csv=None, 
                 proj_name='UnnamedProject'):
        # params from input arguments
        self.in_df = in_df
        self.xl_template = xl_template
        self.import_tab = import_tab
        self.xl_out = xl_out
        self.xlsheets_to_pdf = xlsheets_to_pdf
        self.map_key_csv = map_key_csv  # {<image file name>: [<sheet to put image on>, <row>, <col>]}
        self.proj_name = proj_name
        
        # params that are derived or imported from ppa_input_params.py
        self.xl_workbook = openpyxl.load_workbook(self.xl_template)
        self.time_sufx = str(dt.datetime.now().strftime('%m%d%Y%H%M'))
        self.sheets_all_rpts = params.sheets_all_reports
        self.out_folder = arcpy.env.scratchFolder
        self.xl_out_path = os.path.join(self.out_folder, self.xl_out)


    def overwrite_df_to_xlsx(self, unused=0, start_row=0, start_col=0):  # why does there need to be an argument?
        '''Writes pandas dataframe <in_df_ to <tab_name> sheet of <xlsx_template> excel workbook.'''
        in_df = self.in_df.reset_index()
        df_records = in_df.to_records(index=False)
        
        # get header row for output
        out_header_list = [list(in_df.columns)]  # get header row for output
        
        out_data_list = [list(i) for i in df_records]  # get output data rows
    
        comb_out_list = out_header_list + out_data_list
    
        ws = self.xl_workbook[self.import_tab]
        for i, row in enumerate(comb_out_list):
            for j, val in enumerate(row):
                cell = ws.cell(row=(start_row + (i + 1)), column=(start_col + (j + 1)))
                if (cell):
                    cell.value = val
    
    def insert_image_xlsx(self, sheet_name, rownum, col_letter, img_file):
        '''inserts image into specified sheet and cell within Excel workbook'''
        ws = self.xl_workbook[sheet_name]
        cell = '{}{}'.format(col_letter, rownum) # will be where upper left corner of image placed
        img_obj = Image(img_file)
        ws.add_image(img_obj, cell)
        
    def make_new_excel(self):
        '''takes excel template > writes new values to import/updates charts, then inserts indicated images at specified locations'''
        self.overwrite_df_to_xlsx(self) # write data to import tab
        
        # insert map images at appropriate locations
        # NEED TEMP FOLDER FOR MAP IMAGES TO WRITE TO, WITHIN SCRATCH GDB, THEN CALL IMGS FROM THIS FOLDER
        
        if self.map_key_csv:
            bookname = os.path.basename(self.xl_template)
            mapkey = pd.read_csv(self.map_key_csv)
            mapkey_dict_list = mapkey.loc[mapkey['Report'] == bookname] \
                .to_dict(orient='records') # filter master table to only get tabs for workbook corresponding to specified project type
            
            for i in mapkey_dict_list:
                imgdir = i['MapImgDir']
                imgfile = i['MapImgFile']
                
                if isinstance(imgfile, str):  # if there is an imgfile...
                    imgfilepath = os.path.join(imgdir, imgfile)
                
                    sheet = i['Tab']
                    row = i['RowNum']
                    col = i['ColNum']
                    
                    self.insert_image_xlsx(sheet, row, col, imgfilepath)
        
        self.xl_workbook.save(self.xl_out_path)
        self.xl_workbook.close()
        
    def make_pdf(self):
        wb = None
        out_pdf_final_name = "Rpt{}{}.pdf".format(self.proj_name, self.time_sufx)
        out_pdf_final = os.path.join(self.out_folder, out_pdf_final_name)
            
        if os.path.exists(out_pdf_final):
            try:
                os.remove(out_pdf_final)
            except:
                out_pdf_final_name = "Rpt{}{}v2.pdf".format(self.proj_name, self.time_sufx)
                out_pdf_final = os.path.join(self.out_folder, out_pdf_final_name)
                
        try:
            arcpy.AddMessage("Publishing to PDF...")
            arcpy.AddMessage("Excel output file: {}".format(self.xl_out_path))

            # make excel workbook with project outputs
            xw.App.visible = True # must be set to False or else it won't work
            
            #consider making try/catch
            if not os.path.exists(self.xl_out_path):
                self.make_new_excel()
            
            wb = xw.Book(self.xl_out_path)
            arcpy.AddMessage("sheets in book: {}".format(wb.sheets))
            
            out_sheets = self.sheets_all_rpts + self.xlsheets_to_pdf
            arcpy.AddMessage("output sheets: {}".format(out_sheets))
                
            l_out_pdfs = [] # will be list of output PDF file names, PDFs in this list will be combined. Need list step for sorting by sheet name A-Z
            
            pdf_final = arcpy.mp.PDFDocumentCreate(out_pdf_final) # instantiate arcpy PDFDocumentCreate object
                
            # write user-specified sheets to PDFs
            for s in out_sheets:
                out_sheet = wb.sheets[s]
                pdf_out = os.path.join(self.out_folder, 'Sheet_{}_{}.pdf'.format(s, self.time_sufx))
                out_sheet.api.ExportAsFixedFormat(0, pdf_out)
                l_out_pdfs.append(pdf_out)
            
            l_out_pdfs = sorted(l_out_pdfs) # sort by sheet name so that PDFs append together in correct order.
            
            for singlesheet_pdf in l_out_pdfs:
                pdf_final.appendPages(singlesheet_pdf) # append to master PDF object
                
                try:
                    os.remove(singlesheet_pdf)  # not necessary for online version because scratch folder is temporary
                except:
                    pass
                
            pdf_final.saveAndClose()
            
            
            t_returns = (params.msg_ok, out_pdf_final, self.xl_out_path) # if successful, return output excel and output PDF
            # wb.close()  # if error in the above for loop, then it never reaches this line and wb never closes.
        except:
            msg = "{}".format(trace())
            arcpy.AddMessage(msg)
            
            t_returns = (msg, ) # if fail, return error message
        finally: # always runs, even if 'try' runs successfully.
            if wb != None:  # only closes wb object if it was instantiated.
                wb.close()
            gc.collect()
            arcpy.AddMessage(self.xl_out_path)
            
        return t_returns




